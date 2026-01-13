import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from jobspy import scrape_jobs
from datetime import datetime

# Configuration
RESULTS_PER_SITE = 50  # Can be: 15, 20, 50, 100+ (higher = more results but slower)
SEARCH_TERM = "Data Engineer"
LOCATION = "United Kingdom"
SITES = ["indeed", "linkedin", "glassdoor"]

# Scrape jobs
print(f"🔍 Searching for {SEARCH_TERM} jobs in {LOCATION}...")
print(f"📊 Results per site: {RESULTS_PER_SITE}")

jobs = scrape_jobs(
    site_name=SITES,
    search_term=SEARCH_TERM,
    location=LOCATION,
    results_wanted=RESULTS_PER_SITE,
    country_indeed='UK',
)

print(f"Found {len(jobs)} jobs")

# Generate filename with timestamp
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file = f'data_engineer/data_engineer_jobs_{timestamp}.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    workbook = writer.book
    
    # Define styling
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Sheet 1: Summary Statistics
    summary_data = {
        'Metric': [
            'Total Jobs Found',
            'Unique Job Sites',
            'Unique Companies',
            'Remote Jobs',
            'Average Min Salary (£)',
            'Average Max Salary (£)',
            'Most Common Job Type'
        ],
        'Value': [
            len(jobs),
            jobs['site'].nunique(),
            jobs['company'].nunique(),
            jobs['is_remote'].sum() if 'is_remote' in jobs.columns else 0,
            f"£{jobs['min_amount'].mean():,.2f}" if jobs['min_amount'].notna().any() else "N/A",
            f"£{jobs['max_amount'].mean():,.2f}" if jobs['max_amount'].notna().any() else "N/A",
            jobs['job_type'].mode()[0] if 'job_type' in jobs.columns and len(jobs['job_type'].mode()) > 0 else "N/A"
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Format summary sheet
    summary_sheet = writer.sheets['Summary']
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sheet 2: All Jobs
    jobs_display = jobs[[col for col in [
        'site', 'title', 'company', 'location', 'job_type', 
        'min_amount', 'max_amount', 'interval', 'is_remote',
        'date_posted', 'job_level', 'company_industry'
    ] if col in jobs.columns]].copy()
    
    jobs_display.to_excel(writer, sheet_name='All Jobs', index=False)
    
    # Format all jobs sheet
    all_jobs_sheet = writer.sheets['All Jobs']
    for cell in all_jobs_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in all_jobs_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        all_jobs_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3+: Jobs by Site
    for site in sorted(jobs['site'].unique()):
        site_jobs = jobs[jobs['site'] == site]
        site_sheet_name = site.capitalize()[:31]  # Excel sheet name limit
        
        site_jobs.to_excel(writer, sheet_name=site_sheet_name, index=False)
        
        # Format site-specific sheet
        site_sheet = writer.sheets[site_sheet_name]
        for cell in site_sheet[1]:
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in site_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            site_sheet.column_dimensions[column_letter].width = adjusted_width

print(f"\n✅ Excel file created successfully!")
print(f"📁 File: {output_file}")
print(f"📊 Sheets created:")
print(f"   - Summary (statistics)")
print(f"   - All Jobs ({len(jobs)} jobs)")
for site in sorted(jobs['site'].unique()):
    site_count = len(jobs[jobs['site'] == site])
    print(f"   - {site.capitalize()} ({site_count} jobs)")